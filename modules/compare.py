import os
import ctypes
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook


class FolderCompare:
    def __init__(self, folder_a, folder_b, progress, progress_percent):
        self.folder_a = Path(folder_a)
        self.folder_b = Path(folder_b)

        self.progress = progress
        self.progress_percent = progress_percent

        self.start_time = datetime.now()

        log_name = self.start_time.strftime("%Y%m%d-%H%M%S")

        self.report_file = Path("logs/Compare")/ f"{log_name}.xlsx"

        self.a_total_files = 0
        self.a_total_folders = 0
        self.a_hidden_files = 0
        self.a_hidden_folders = 0
        self.a_total_size = 0

        self.b_total_files = 0
        self.b_total_folders = 0
        self.b_hidden_files = 0
        self.b_hidden_folders = 0
        self.b_total_size = 0

        self.a_files = {}
        self.b_files = {}

        self.a_folders = set()
        self.b_folders = set()

    def is_hidden(self, path):

        try:
            attrs = ctypes.windll.kernel32.GetFileAttributesW(
                str(path)
            )

            return bool(attrs & 2)

        except:
            return False

    def format_size(self, size):

        for unit in ["B", "KB", "MB", "GB", "TB"]:

            if size < 1024:
                return f"{size:.2f} {unit}"

            size /= 1024

        return f"{size:.2f} PB"

    def scan_folder(self, folder, folder_name):
        files_dict = {}
        folders_set = set()

        total_files = 0
        total_folders = 0
        hidden_files = 0
        hidden_folders = 0
        total_size = 0
        for root, dirs, files in os.walk(folder):
            for dir_name in dirs:

                folder_path = Path(root) / dir_name

                relative_path = str(
                    folder_path.relative_to(folder)
                )

                folders_set.add(relative_path)

                total_folders += 1

                if self.is_hidden(folder_path):
                    hidden_folders += 1

            for file_name in files:

                file_path = Path(root) / file_name

                relative_path = str(
                    file_path.relative_to(folder)                    )

                try:

                    size = file_path.stat().st_size

                    files_dict[relative_path] = size

                    total_files += 1
                    total_size += size

                    if self.is_hidden(file_path):
                        hidden_files += 1

                except:
                    continue

        if folder_name == "A":

            self.a_files = files_dict
            self.a_folders = folders_set

            self.a_total_files = total_files
            self.a_total_folders = total_folders
            self.a_hidden_files = hidden_files
            self.a_hidden_folders = hidden_folders
            self.a_total_size = total_size

            self.progress.after(
                0,
                lambda: self.progress.set(0.25)
            )

            self.progress_percent.after(
                0,
                lambda: self.progress_percent.configure(
                    text="25%"
                )
            )

        else:

            self.b_files = files_dict
            self.b_folders = folders_set

            self.b_total_files = total_files
            self.b_total_folders = total_folders
            self.b_hidden_files = hidden_files
            self.b_hidden_folders = hidden_folders
            self.b_total_size = total_size

            self.progress.after(
                0,
                lambda: self.progress.set(0.50)
            )

            self.progress_percent.after(
                0,
                lambda: self.progress_percent.configure(
                    text="50%"
                )
            )

    def compare_folders(self):

        self.missing_in_b = []
        self.extra_in_b = []
        self.modified_files = []

        self.missing_folders_in_b = []
        self.extra_folders_in_b = []

        #Missing_Files
        for file_path in self.a_files:

            if file_path not in self.b_files:
                self.missing_in_b.append(
                    file_path
                )

        #Extra_Files
        for file_path in self.b_files:

            if file_path not in self.a_files:
                self.extra_in_b.append(
                    file_path
                )

        #Modified_Files
        for file_path in self.a_files:

            if file_path in self.b_files:

                if self.a_files[file_path] != self.b_files[file_path]:
                    self.modified_files.append(
                        (
                            file_path,
                            self.a_files[file_path],
                            self.b_files[file_path]
                        )
                    )

        #Missing_Folders
        for folder in self.a_folders:

            if folder not in self.b_folders:
                self.missing_folders_in_b.append(
                    folder
                )

        #Extra_Folders
        for folder in self.b_folders:

            if folder not in self.a_folders:
                self.extra_folders_in_b.append(
                    folder
                )

        self.progress.after(
            0,
            lambda: self.progress.set(0.75)
        )

        self.progress_percent.after(
            0,
            lambda: self.progress_percent.configure(
                text="75%"
            )
        )

    def create_excel_report(self):

        workbook = Workbook()

        # =========================
        # Summary Sheet
        # =========================

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        summary_sheet.append(
            ["Folder A", str(self.folder_a)]
        )

        summary_sheet.append(
            ["Folder B", str(self.folder_b)]
        )

        summary_sheet.append([])

        summary_sheet.append(
            ["Property", "Folder A", "Folder B"]
        )

        summary_sheet.append(
            [
                "Total Files",
                self.a_total_files,
                self.b_total_files
            ]
        )

        summary_sheet.append(
            [
                "Total Folders",
                self.a_total_folders,
                self.b_total_folders
            ]
        )

        summary_sheet.append(
            [
                "Hidden Files",
                self.a_hidden_files,
                self.b_hidden_files
            ]
        )

        summary_sheet.append(
            [
                "Hidden Folders",
                self.a_hidden_folders,
                self.b_hidden_folders
            ]
        )

        summary_sheet.append(
            [
                "Total Size",
                self.format_size(self.a_total_size),
                self.format_size(self.b_total_size)
            ]
        )

        #Missing_File_Sheet
        missing_sheet = workbook.create_sheet(
            "Missing Files"
        )

        missing_sheet.append(
            ["Present in A but Missing in B"]
        )

        for file_path in self.missing_in_b:
            missing_sheet.append(
                [file_path]
            )

        #Extra_File_Sheet
        extra_sheet = workbook.create_sheet(
            "Extra Files"
        )

        extra_sheet.append(
            ["Present in B but Missing in A"]
        )

        for file_path in self.extra_in_b:
            extra_sheet.append(
                [file_path]
            )

        #Modified_File_Sheet
        modified_sheet = workbook.create_sheet(
            "Modified Files"
        )

        modified_sheet.append(
            [
                "File Path",
                "Folder A Size",
                "Folder B Size"
            ]
        )

        for file_path, a_size, b_size in self.modified_files:
            modified_sheet.append(
                [
                    file_path,
                    self.format_size(a_size),
                    self.format_size(b_size)
                ]
            )

        #Folder_Difference_Sheet
        folder_sheet = workbook.create_sheet(
            "Folder Differences"
        )

        folder_sheet.append(
            [
                "Missing In B",
                "Extra In B"
            ]
        )

        max_rows = max(
            len(self.missing_folders_in_b),
            len(self.extra_folders_in_b)
        )

        for i in range(max_rows):

            missing = ""

            extra = ""

            if i < len(self.missing_folders_in_b):
                missing = self.missing_folders_in_b[i]

            if i < len(self.extra_folders_in_b):
                extra = self.extra_folders_in_b[i]

            folder_sheet.append(
                [
                    missing,
                    extra
                ]
            )

        #Auto_Size_Columns
        for sheet in workbook.worksheets:

            for column in sheet.columns:

                max_length = 0

                column_letter = column[0].column_letter

                for cell in column:

                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(
                                str(cell.value)
                            )

                    except:
                        pass

                sheet.column_dimensions[
                    column_letter
                ].width = max_length + 5

        # Save_Reports
        workbook.save(
            self.report_file
        )

        #Update_Progress
        self.progress.after(
            0,
            lambda: self.progress.set(1)
        )

        self.progress_percent.after(
            0,
            lambda: self.progress_percent.configure(
                text="100%"
            )
        )

    #Main_Execute_Function
    def run(self):

        self.scan_folder(
            self.folder_a,
            "A"
        )

        self.scan_folder(
            self.folder_b,
            "B"
        )

        self.compare_folders()

        self.create_excel_report()

        return self.report_file